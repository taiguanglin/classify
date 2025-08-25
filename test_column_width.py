#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
測試Excel列寬設置的腳本
"""

import openpyxl
import sys

def test_column_widths(excel_file):
    """測試Excel文件的列寬設置"""
    try:
        print(f"📊 檢查Excel文件: {excel_file}")
        print("=" * 60)
        
        # 載入Excel文件
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        print(f"📋 工作表名稱: {worksheet.title}")
        print(f"📏 總行數: {worksheet.max_row}")
        print(f"📏 總列數: {worksheet.max_column}")
        print()
        
        # 檢查評分相關列的寬度
        scoring_columns = [
            (24, "廣度評分"),
            (25, "深度評分"),
            (26, "獨特性評分"),
            (27, "綜合評分"),
            (28, "廣度評論"),
            (29, "深度評論"),
            (30, "獨特性評論"),
            (31, "總體評價")
        ]
        
        print("🔍 評分相關列寬度檢查:")
        print("-" * 40)
        
        for col_num, col_name in scoring_columns:
            if col_num <= worksheet.max_column:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                col_width = worksheet.column_dimensions[col_letter].width
                
                # 檢查列標題
                header_cell = worksheet.cell(row=1, column=col_num)
                header_value = header_cell.value if header_cell.value else "無標題"
                
                print(f"列 {col_letter} ({col_num}): {col_name}")
                print(f"  標題: {header_value}")
                print(f"  寬度: {col_width}")
                
                # 檢查內容長度
                max_content_length = 0
                sample_content = ""
                # 檢查特定的行（根據結果文件中的行號）
                specific_rows = [82, 86, 292, 332, 405, 463, 464, 466, 481, 492, 512]
                for row in specific_rows:
                    if row <= worksheet.max_row:
                        cell = worksheet.cell(row=row, column=col_num)
                        if cell.value:
                            content = str(cell.value)
                            if len(content) > max_content_length:
                                max_content_length = len(content)
                                sample_content = content[:100] + "..." if len(content) > 100 else content
                
                print(f"  最大內容長度: {max_content_length}")
                print(f"  樣本內容: {sample_content}")
                print()
            else:
                print(f"列 {col_num}: 超出範圍")
                print()
        
        # 檢查問題和答案列
        print("🔍 問題和答案列檢查:")
        print("-" * 40)
        
        question_col = 18  # R列
        answer_col = 19    # S列
        
        # 檢查特定行的評論
        specific_rows = [82, 86, 292, 332, 405, 463, 464, 466, 481, 492, 512]
        
        for col_num, col_name in [(question_col, "問題"), (answer_col, "答案")]:
            if col_num <= worksheet.max_column:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                col_width = worksheet.column_dimensions[col_letter].width
                
                print(f"列 {col_letter} ({col_num}): {col_name}")
                print(f"  寬度: {col_width}")
                
                # 檢查是否有評論
                comment_found = False
                for row in specific_rows:
                    if row <= worksheet.max_row:
                        cell = worksheet.cell(row=row, column=col_num)
                        if cell.comment:
                            comment_found = True
                            print(f"  行{row}有評論: 是")
                            print(f"  評論內容: {cell.comment.text[:100]}...")
                            break
                
                if not comment_found:
                    print(f"  在檢查的行中沒有找到評論")
                print()
        
        workbook.close()
        print("✅ 列寬檢查完成")
        
    except Exception as e:
        print(f"❌ 檢查失敗: {e}")
        return False
    
    return True

def main():
    """主函數"""
    if len(sys.argv) != 2:
        print("使用方法: python3 test_column_width.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    test_column_widths(excel_file)

if __name__ == "__main__":
    main()
