#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
問答分類器 v2.0 - 輸出到JSON文件版本
將分類處理與Excel寫入分離，提高效能和容錯性
"""

import configparser
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openai import OpenAI
import re
import time
import logging
from datetime import datetime
import os
import json
import argparse
from typing import Dict, List, Tuple, Optional

# 設置日誌函數
def setup_logging():
    """設置日誌配置"""
    # 確保日誌文件存在
    log_file = "qa_classification.log"
    if not os.path.exists(log_file):
        with open(log_file, "w") as f:
            f.write(f"# QA分類器日誌文件 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n")
    
    # 創建自定義logger
    logger = logging.getLogger("qa_classifier")
    logger.setLevel(logging.INFO)
    
    # 清除現有handlers
    if logger.handlers:
        logger.handlers.clear()
    
    # 創建文件handler
    file_handler = logging.FileHandler(log_file, mode="a", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    
    # 創建控制台handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # 設置格式
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # 添加handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    logger.info(f"日誌系統初始化完成 - 日誌文件: {log_file}")
    
    return logger


# 初始化日誌系統
logger = setup_logging()

class QAClassifierV2:
    """問答分類器 v2.0 - JSON輸出版本"""
    
    def __init__(self, config_file: str = 'config.ini', api_key: str = None, api_type: str = None, chatmock_url: str = None):
        """初始化分類器"""
        self.config = configparser.ConfigParser()
        self.config.read(config_file, encoding='utf-8')
        
        # 保存参数
        self.api_key = api_key
        self.api_type = api_type
        self.chatmock_url = chatmock_url
        
        # 初始化OpenAI
        self.setup_openai()
        
        # 載入分類體系
        self.category_system = self.load_category_system()
        
        # 載入prompt模板
        self.prompt_template = self.load_prompt_template()
        
        # 結果存儲
        self.results = {}
        self.metadata = {
            "source_file": self.config.get('excel', 'file_path'),
            "sheet_name": self.config.get('excel', 'sheet_name'),
            "llm_model": self._get_llm_model_display_name(),  # 动态获取模型显示名称
            "processing_start_time": datetime.now().isoformat(),
            "total_processed": 0,
            "total_success": 0,
            "total_failed": 0
        }
        
        logger.info("QA分類器 v2.0 初始化完成")
    
    def setup_openai(self):
        """設置OpenAI API或ChatMock"""
        # 優先使用命令行參數，其次使用配置文件
        if self.api_type:
            api_type = self.api_type.lower()
        else:
            api_type = self.config.get('api', 'type', fallback='openai').lower()
        
        logger.info(f"使用API類型: {api_type}")
        
        if api_type == 'chatmock':
            self._setup_chatmock()
        else:
            self._setup_openai_api()
    
    def _setup_chatmock(self):
        """設置ChatMock本地服務器"""
        try:
            # 優先使用命令行參數，其次使用配置文件
            if self.chatmock_url:
                base_url = self.chatmock_url
            else:
                base_url = self.config.get('chatmock', 'base_url', fallback='http://127.0.0.1:8000/v1')
            
            model = self.config.get('chatmock', 'model', fallback='gpt-5')
            
            # 創建OpenAI客戶端實例，指向ChatMock服務器
            self.client = OpenAI(
                base_url=base_url,
                api_key="chatmock"  # ChatMock忽略此值
            )
            self.model = model
            
            # ChatMock使用GPT-5參數
            self.temperature, self.max_tokens = self._get_model_specific_params()
            
            logger.info(f"ChatMock設置完成 - 服務器: {base_url}")
            logger.info(f"使用模型: {self.model}")
            logger.info(f"使用參數 - 溫度: {self.temperature}, 最大Token: {self.max_tokens}")
            
        except Exception as e:
            logger.error(f"ChatMock設置失敗: {e}")
            raise ValueError(f"ChatMock設置失敗: {e}")
    
    def _setup_openai_api(self):
        """設置OpenAI官方API"""
        # 優先使用傳入的API key
        if self.api_key:
            api_key = self.api_key
        else:
            # 嘗試從環境變量獲取
            api_key = os.getenv('OPENAI_API_KEY')
            if not api_key:
                # 最後嘗試從配置文件獲取（向後兼容）
                api_key = self.config.get('openai', 'api_key', fallback=None)
        
        if not api_key or api_key == 'YOUR_OPENAI_API_KEY_HERE':
            raise ValueError(
                "請通過以下方式之一設置OpenAI API Key:\n"
                "1. 命令行參數: --api-key YOUR_API_KEY\n"
                "2. 環境變量: export OPENAI_API_KEY=YOUR_API_KEY\n"
                "3. 配置文件: 在config.ini中設置api_key（不推薦）"
            )
        
        # 創建OpenAI客戶端實例
        self.client = OpenAI(api_key=api_key)
        self.model = self.config.get('openai', 'model', fallback='gpt-4')
        
        # 根據模型類型自動選擇參數配置
        self.temperature, self.max_tokens = self._get_model_specific_params()
        
        logger.info(f"OpenAI設置完成 - 模型: {self.model}")
        logger.info(f"使用參數 - 溫度: {self.temperature}, 最大Token: {self.max_tokens}")
    
    def _get_model_specific_params(self) -> tuple:
        """根據模型類型獲取對應的參數配置"""
        try:
            if self.model.startswith('gpt-5'):
                # GPT-5系列模型參數
                
                # GPT-5不支持自定義temperature，使用默認值1
                temperature = 1.0
                
                # 嘗試讀取max_completion_tokens，但不強制要求
                try:
                    max_tokens = self.config.getint('gpt5_models', 'max_completion_tokens', fallback=None)
                    if max_tokens is not None:
                        logger.warning("檢測到max_completion_tokens設置，但建議不設置以避免空回應")
                except:
                    max_tokens = None
                
                logger.info(f"使用GPT-5專用參數配置")
                
            else:
                # GPT-4系列模型參數
                logger.info("使用GPT-4專用參數配置")
                
                temperature = self.config.getfloat('gpt4_models', 'temperature', fallback=0.3)
                max_tokens = self.config.getint('gpt4_models', 'max_tokens', fallback=1000)
            
            return temperature, max_tokens
            
        except Exception as e:
            logger.error(f"獲取模型特定參數失敗: {e}")
            logger.warning("使用默認參數配置")
            return 0.3, 1000
    
    def _get_llm_model_display_name(self) -> str:
        """獲取LLM模型的顯示名稱，根據API類型動態設置"""
        try:
            # 獲取當前API類型
            if self.api_type:
                api_type = self.api_type.lower()
            else:
                api_type = self.config.get('api', 'type', fallback='openai').lower()
            
            if api_type == 'chatmock':
                # ChatMock模式：使用ChatMock配置的模型名稱
                model = self.config.get('chatmock', 'model', fallback='gpt-5')
                return f"chat-{model}"  # 例如：chat-gpt-5
            else:
                # OpenAI模式：使用OpenAI配置的模型名稱
                model = self.config.get('openai', 'model', fallback='gpt-4')
                return model  # 例如：gpt-5-nano
            
        except Exception as e:
            logger.error(f"獲取模型顯示名稱失敗: {e}")
            # 返回默認值
            return "gpt-4"
    
    def load_category_system(self) -> str:
        """載入新目錄分類體系"""
        toc_file = "/Users/paul/taiguanglin.github.io/new_toc/wenda2-toc-2025-08-18v2.txt"
        
        if not os.path.exists(toc_file):
            logger.warning(f"目錄文件不存在: {toc_file}")
            return "目錄體系載入失敗"
        
        try:
            with open(toc_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 提取主要分類和次級分類
            categories = []
            lines = content.strip().split('\n')
            
            for line in lines:
                if line.strip():
                    # 使用正則表達式提取分類名稱
                    match = re.search(r'([^\(\d]+)', line.strip())
                    if match:
                        category = match.group(1).strip()
                        if len(category) > 2:  # 過濾太短的分類名
                            categories.append(category)
            
            # 去重並格式化
            unique_categories = list(set(categories))
            category_text = "、".join(unique_categories[:20])  # 限制數量避免prompt太長
            
            logger.info(f"成功載入 {len(unique_categories)} 個分類")
            return category_text
            
        except Exception as e:
            logger.error(f"載入目錄體系失敗: {e}")
            return "目錄體系載入失敗"
    
    def load_prompt_template(self) -> str:
        """載入prompt模板"""
        # 优先使用简化的prompt模板
        prompt_file = 'prompt_template_simple.txt'
        
        if not os.path.exists(prompt_file):
            # 如果简化模板不存在，使用原始模板
            prompt_file = 'prompt_template.txt'
            if not os.path.exists(prompt_file):
                logger.warning(f"Prompt文件不存在: {prompt_file}")
                return self.get_default_prompt()
        
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"載入prompt模板失敗: {e}")
            return self.get_default_prompt()
    
    def get_default_prompt(self) -> str:
        """獲取默認prompt模板"""
        return """請根據以下目錄體系對問答進行分類：

目錄體系：
{category_system}

問答標題：{title}
問答內容：{qa_content}

請按以下格式回答：

✅ 最佳分類排序：
【分類名稱】（置信度%）【分類名稱】（置信度%）【分類名稱】（置信度%）

✅ 理由：
(說明分類理由，150字以內)

✅ 提問重點摘要：
(提問的重點與摘要，100字以內)

✅ 回答重點摘要：
(回答的重點與摘要，150字以內)"""

    def load_excel_data(self) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
        """載入Excel數據"""
        file_path = self.config.get('excel', 'file_path')
        sheet_name = self.config.get('excel', 'sheet_name')
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook[sheet_name]
            logger.info(f"成功載入Excel文件: {file_path}, 工作表: {sheet_name}")
            return workbook, worksheet
        except Exception as e:
            logger.error(f"載入Excel文件失敗: {e}")
            raise

    def extract_qa_content(self, worksheet, row: int) -> Tuple[str, str]:
        """提取問答內容"""
        # 根據Excel結構調整列號
        question_col = 8  # 第8列是問題
        answer_col = 9    # 第9列是答案
        
        try:
            question = worksheet.cell(row=row, column=question_col).value or ""
            answer = worksheet.cell(row=row, column=answer_col).value or ""
            
            return str(question).strip(), str(answer).strip()
        except Exception as e:
            logger.error(f"提取第 {row} 行內容失敗: {e}")
            return "", ""

    def classify_qa(self, title: str, content: str) -> Dict[str, str]:
        """使用OpenAI對問答進行分類"""
        if not title and not content:
            return {
                'classification': '無法分類',
                'reason': '標題和內容均為空',
                'question_summary': '',
                'answer_summary': '',
                'status': 'empty'
            }
        
        # 構建prompt
        prompt = self.prompt_template.format(
            category_system=self.category_system,
            title=title
        )
        
        # 记录发送给LLM的完整prompt（调试时使用）
        # logger.info(f"发送给LLM的完整prompt: {prompt}")
        
        try:
            # 根据模型类型选择正确的参数
            api_params = {
                "model": self.model,
                "messages": [
                    {"role": "system", "content": "你是一個專業的佛學問答分類專家。"},
                    {"role": "user", "content": prompt}
                ]
            }
            
            # 检查是否使用ChatMock（与初始化时保持一致）
            if self.api_type:
                api_type = self.api_type.lower()
            else:
                api_type = self.config.get('api', 'type', fallback='openai').lower()
            
            if api_type == 'chatmock':
                # ChatMock特有的参数
                reasoning_effort = self.config.get('chatmock', 'reasoning_effort', fallback='medium')
                reasoning_summary = self.config.get('chatmock', 'reasoning_summary', fallback='auto')
                
                if reasoning_effort and reasoning_effort != 'medium':
                    api_params["reasoning_effort"] = reasoning_effort
                if reasoning_summary and reasoning_summary != 'auto':
                    api_params["reasoning_summary"] = reasoning_summary
                
                logger.debug(f"ChatMock参数 - 推理努力: {reasoning_effort}, 推理摘要: {reasoning_summary}")
            else:
                # OpenAI API参数
                if self.model.startswith('gpt-5'):
                    # GPT-5模型不支持自定义temperature，也不设置max_completion_tokens
                    # 不设置temperature参数，使用默认值
                    # 不设置max_completion_tokens，让模型使用默认限制
                    logger.debug("GPT-5模型：使用默认temperature和token限制")
                else:
                    # 其他模型使用max_tokens和temperature
                    api_params["temperature"] = self.temperature
                    api_params["max_tokens"] = self.max_tokens
                    logger.debug(f"GPT-4模型：使用temperature={self.temperature}, max_tokens={self.max_tokens}")
            
            response = self.client.chat.completions.create(**api_params)
            
            result_text = response.choices[0].message.content.strip()
            
            # 保存原始响应用于调试（调试时使用）
            # logger.info(f"原始API响应: {result_text}")
            
            parsed_result = self.parse_classification_result(result_text)
            parsed_result['status'] = 'success'
            
            # 保存原始响应到结果中
            parsed_result['raw_response'] = result_text
            
            return parsed_result
            
        except Exception as e:
            logger.error(f"OpenAI API調用失敗: {e}")
            return {
                'classification': 'API錯誤',
                'reason': f'API調用失敗: {str(e)}',
                'question_summary': '',
                'answer_summary': '',
                'status': 'api_error'
            }

    def parse_classification_result(self, result_text: str) -> Dict[str, str]:
        """解析分類結果"""
        try:
            # 使用正則表達式提取各部分
            classification_match = re.search(r'✅ 最佳分類排序：\s*\n(.+?)(?=\n\n|✅|$)', result_text, re.DOTALL)
            reason_match = re.search(r'✅ 理由：\s*\n(.+?)(?=\n\n|✅|$)', result_text, re.DOTALL)
            question_summary_match = re.search(r'✅ 提問重點摘要：\s*\n(.+?)(?=\n\n|✅|$)', result_text, re.DOTALL)
            answer_summary_match = re.search(r'✅ 回答重點摘要：\s*\n(.+?)(?=\n\n|✅|$)', result_text, re.DOTALL)
            
            return {
                'classification': classification_match.group(1).strip() if classification_match else '解析失敗',
                'reason': reason_match.group(1).strip() if reason_match else '解析失敗',
                'question_summary': question_summary_match.group(1).strip() if question_summary_match else '解析失敗',
                'answer_summary': answer_summary_match.group(1).strip() if answer_summary_match else '解析失敗'
            }
        except Exception as e:
            logger.error(f"解析分類結果失敗: {e}")
            return {
                'classification': '解析錯誤',
                'reason': f'解析失敗: {str(e)}',
                'question_summary': '解析錯誤',
                'answer_summary': '解析錯誤'
            }

    def load_existing_results(self, results_file: str) -> Dict:
        """載入已有的分類結果"""
        if not os.path.exists(results_file):
            return {}
        
        try:
            with open(results_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('results', {})
        except Exception as e:
            logger.error(f"載入已有結果失敗: {e}")
            return {}

    def save_results(self, results_file: str):
        """保存分類結果到JSON文件"""
        try:
            self.metadata['processing_end_time'] = datetime.now().isoformat()
            self.metadata['total_processed'] = len(self.results)
            self.metadata['total_success'] = sum(1 for r in self.results.values() if r.get('status') == 'success')
            self.metadata['total_failed'] = len(self.results) - self.metadata['total_success']
            
            output_data = {
                'metadata': self.metadata,
                'results': self.results
            }
            
            with open(results_file, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"結果已保存到: {results_file}")
            logger.info(f"總處理: {self.metadata['total_processed']}, 成功: {self.metadata['total_success']}, 失敗: {self.metadata['total_failed']}")
            
        except Exception as e:
            logger.error(f"保存結果失敗: {e}")

    def process_batch(self, start_row: int = None, end_row: int = None, results_file: str = None):
        """批量處理問答分類，輸出到JSON文件"""
        # 載入配置
        if start_row is None:
            start_row = self.config.getint('processing', 'start_row', fallback=2)
        if end_row is None:
            config_end_row = self.config.getint('processing', 'end_row', fallback=0)
            end_row = config_end_row if config_end_row > 0 else None
        
        if results_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            results_file = f'qa_classification_results_{timestamp}.json'
        
        # 載入已有結果（支持續處理）
        self.results = self.load_existing_results(results_file)
        
        # 載入Excel
        workbook, worksheet = self.load_excel_data()
        
        # 確定處理範圍
        max_row = worksheet.max_row
        if end_row is None or end_row > max_row:
            end_row = max_row
        
        logger.info(f"開始處理第 {start_row} 到 {end_row} 行，共 {end_row - start_row + 1} 條記錄")
        logger.info(f"結果將保存到: {results_file}")
        
        processed_count = 0
        success_count = 0
        
        for row in range(start_row, end_row + 1):
            try:
                # 檢查是否已處理
                row_key = str(row)
                if row_key in self.results:
                    logger.info(f"第 {row} 行已處理，跳過")
                    continue
                
                # 提取問答內容
                question, answer = self.extract_qa_content(worksheet, row)
                
                if not question and not answer:
                    logger.info(f"第 {row} 行無內容，跳過")
                    continue
                
                logger.info(f"處理第 {row} 行: {question[:50]}...")
                
                # 進行分類
                result = self.classify_qa(question, answer)
                
                # 保存結果
                self.results[row_key] = {
                    'row_number': row,
                    'question': question[:500],  # 限制長度
                    'answer': answer[:1000],     # 限制長度
                    'classification': result['classification'],
                    'reason': result['reason'],
                    'question_summary': result['question_summary'],
                    'answer_summary': result['answer_summary'],
                    'status': result['status'],
                    'processed_time': datetime.now().isoformat()
                }
                
                processed_count += 1
                if result['status'] == 'success':
                    success_count += 1
                
                logger.info(f"第 {row} 行處理完成")
                
                # 每處理10條記錄保存一次
                if processed_count % 10 == 0:
                    self.save_results(results_file)
                    logger.info(f"已處理 {processed_count} 條記錄，中間保存完成")
                
                # API調用間隔
                time.sleep(1)
                
            except Exception as e:
                logger.error(f"處理第 {row} 行時發生錯誤: {e}")
                continue
        
        # 最終保存
        self.save_results(results_file)
        
        logger.info(f"批量處理完成！總共處理 {processed_count} 條記錄，成功 {success_count} 條")
        return results_file

def main():
    """主函數"""
    parser = argparse.ArgumentParser(
        description="問答分類自動化系統 v2.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 使用OpenAI API（推薦）
  python3 qa_classifier_v2.py --api-key YOUR_API_KEY --api-type openai
  
  # 使用ChatMock
  python3 qa_classifier_v2.py --api-type chatmock
  
  # 使用環境變量
  export OPENAI_API_KEY=YOUR_API_KEY
  python3 qa_classifier_v2.py --api-type openai
  
  # 使用配置文件（不推薦，會產生commit警告）
  python3 qa_classifier_v2.py
        """
    )
    
    parser.add_argument(
        '--api-key',
        type=str,
        help='OpenAI API Key（推薦使用此方式）'
    )
    
    parser.add_argument(
        '--api-type',
        type=str,
        choices=['openai', 'chatmock'],
        help='API類型選擇：openai 或 chatmock（覆蓋配置文件設置）'
    )
    
    parser.add_argument(
        '--chatmock-url',
        type=str,
        help='ChatMock服務器地址（覆蓋配置文件設置）'
    )
    
    parser.add_argument(
        '--config',
        type=str,
        default='config.ini',
        help='配置文件路徑（默認: config.ini）'
    )
    
    args = parser.parse_args()
    
    print("問答分類自動化系統 v2.0")
    print("=" * 50)
    
    try:
        classifier = QAClassifierV2(
            config_file=args.config,
            api_key=args.api_key,
            api_type=args.api_type,
            chatmock_url=args.chatmock_url
        )
        
        # 處理指定範圍
        results_file = classifier.process_batch()
        
        print(f"\n✅ 分類完成！結果已保存到: {results_file}")
        print("接下來請運行 results_to_excel.py 將結果寫入Excel文件")
        
    except Exception as e:
        logger.error(f"程序執行失敗: {e}")
        print(f"❌ 程序執行失敗: {e}")

if __name__ == "__main__":
    main()
