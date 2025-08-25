#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
佛學問答精選器 - 專門用於評選高質量的佛學問答
支持兩種評分模式：指定行號模式和過濾結果模式
"""

import configparser
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openai import OpenAI
import re
import time
import logging
from datetime import datetime
import os
import json
import argparse
from typing import Dict, List, Tuple, Optional, Union, Any

# 導入緩存系統
try:
    from filter_cache import FilterCache
    FILTER_CACHE_AVAILABLE = True
except ImportError:
    FILTER_CACHE_AVAILABLE = False
    logger = logging.getLogger("qa_curator")
    if logger:
        logger.warning("緩存系統不可用，將使用傳統掃描模式")

# 設置日誌函數
def setup_logging():
    """設置日誌配置"""
    # 確保日誌文件存在
    log_file = "qa_curation.log"
    if not os.path.exists(log_file):
        with open(log_file, "w") as f:
            f.write(f"# 佛學問答精選器日誌文件 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n")
    
    # 創建自定義logger
    logger = logging.getLogger("qa_curator")
    logger.setLevel(logging.INFO)
    
    # 清除現有handlers
    if logger.handlers:
        logger.handlers.clear()
    
    # 創建文件handler
    file_handler = logging.FileHandler(log_file, mode="a", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    
    # 創建控制台handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # 設置格式
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # 添加handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    logger.info(f"日誌系統初始化完成 - 日誌文件: {log_file}")
    
    return logger


# 初始化日誌系統
logger = setup_logging()

class BuddhistQACurator:
    """佛學問答精選器 - 專門用於評選高質量的佛學問答"""
    
    def __init__(self, config_file: str = 'config.ini', api_key: str = None, api_type: str = None, chatmock_url: str = None):
        """初始化精選器"""
        self.config = configparser.ConfigParser()
        self.config.read(config_file, encoding='utf-8')
        
        # 保存参数
        self.api_key = api_key
        self.api_type = api_type
        self.chatmock_url = chatmock_url
        
        # 初始化OpenAI
        self.setup_openai()
        
        # 載入prompt模板
        self.prompt_template = self.load_prompt_template()
        
        # 初始化緩存系統
        if FILTER_CACHE_AVAILABLE:
            cache_dir = self.config.get('filter', 'cache_dir', fallback='.filter_cache')
            self.filter_cache = FilterCache(cache_dir)
            logger.info(f"緩存系統初始化完成，緩存目錄: {cache_dir}")
        else:
            self.filter_cache = None
            logger.warning("緩存系統不可用，將使用傳統掃描模式")
        
        # 結果存儲
        self.curation_results = {}
        self.processing_metadata = {
            "source_file": self.config.get('excel', 'file_path'),
            "sheet_name": self.config.get('excel', 'sheet_name'),
            "llm_model": self._get_llm_model_display_name(),
            "processing_start_time": datetime.now().isoformat(),
            "total_processed": 0,
            "total_success": 0,
            "total_failed": 0,
            "processing_mode": "unknown"
        }
        
        logger.info("佛學問答精選器初始化完成")
    
    def setup_openai(self):
        """設置OpenAI API或ChatMock"""
        # 優先使用命令行參數，其次使用配置文件
        if self.api_type:
            api_type = self.api_type.lower()
        else:
            api_type = self.config.get('api', 'type', fallback='openai').lower()
        
        logger.info(f"使用API類型: {api_type}")
        
        if api_type == 'chatmock':
            self._setup_chatmock()
        else:
            self._setup_openai_api()
    
    def _setup_chatmock(self):
        """設置ChatMock本地服務器"""
        try:
            # 優先使用命令行參數，其次使用配置文件
            if self.chatmock_url:
                base_url = self.chatmock_url
            else:
                base_url = self.config.get('chatmock', 'base_url', fallback='http://127.0.0.1:8000/v1')
            
            model = self.config.get('chatmock', 'model', fallback='gpt-5')
            
            # 創建OpenAI客戶端實例，指向ChatMock服務器
            self.client = OpenAI(
                base_url=base_url,
                api_key="chatmock"  # ChatMock忽略此值
            )
            self.model = model
            
            # ChatMock使用GPT-5參數
            self.temperature, self.max_tokens = self._get_model_specific_params()
            
            logger.info(f"ChatMock設置完成 - 服務器: {base_url}")
            logger.info(f"使用模型: {self.model}")
            logger.info(f"使用參數 - 溫度: {self.temperature}, 最大Token: {self.max_tokens}")
            
        except Exception as e:
            logger.error(f"ChatMock設置失敗: {e}")
            raise ValueError(f"ChatMock設置失敗: {e}")
    
    def _setup_openai_api(self):
        """設置OpenAI官方API"""
        # 優先使用傳入的API key
        if self.api_key:
            api_key = self.api_key
        else:
            # 嘗試從環境變量獲取
            api_key = os.getenv('OPENAI_API_KEY')
            if not api_key:
                # 最後嘗試從配置文件獲取（向後兼容）
                api_key = self.config.get('openai', 'api_key', fallback=None)
        
        if not api_key or api_key == 'YOUR_OPENAI_API_KEY_HERE':
            raise ValueError(
                "請通過以下方式之一設置OpenAI API Key:\n"
                "1. 命令行參數: --api-key YOUR_API_KEY\n"
                "2. 環境變量: export OPENAI_API_KEY=YOUR_API_KEY\n"
                "3. 配置文件: 在config.ini中設置api_key（不推薦）"
            )
        
        # 創建OpenAI客戶端實例
        self.client = OpenAI(api_key=api_key)
        self.model = self.config.get('openai', 'model', fallback='gpt-4')
        
        # 根據模型類型自動選擇參數配置
        self.temperature, self.max_tokens = self._get_model_specific_params()
        
        logger.info(f"OpenAI設置完成 - 模型: {self.model}")
        logger.info(f"使用參數 - 溫度: {self.temperature}, 最大Token: {self.max_tokens}")
    
    def _get_model_specific_params(self) -> tuple:
        """根據模型類型獲取對應的參數配置"""
        try:
            if self.model.startswith('gpt-5'):
                # GPT-5系列模型參數
                
                # GPT-5不支持自定義temperature，使用默認值1
                temperature = 1.0
                
                # 嘗試讀取max_completion_tokens，但不強制要求
                try:
                    max_tokens = self.config.getint('gpt5_models', 'max_completion_tokens', fallback=None)
                    if max_tokens is not None:
                        logger.warning("檢測到max_completion_tokens設置，但建議不設置以避免空回應")
                except:
                    max_tokens = None
                
                logger.info(f"使用GPT-5專用參數配置")
                
            else:
                # GPT-4系列模型參數
                logger.info("使用GPT-4專用參數配置")
                
                temperature = self.config.getfloat('gpt4_models', 'temperature', fallback=0.3)
                max_tokens = self.config.getint('gpt4_models', 'max_tokens', fallback=1000)
            
            return temperature, max_tokens
            
        except Exception as e:
            logger.error(f"獲取模型特定參數失敗: {e}")
            logger.warning("使用默認參數配置")
            return 0.3, 1000
    
    def _get_llm_model_display_name(self) -> str:
        """獲取LLM模型的顯示名稱，根據API類型動態設置"""
        try:
            # 獲取當前API類型
            if self.api_type:
                api_type = self.api_type.lower()
            else:
                api_type = self.config.get('api', 'type', fallback='openai').lower()
            
            if api_type == 'chatmock':
                # ChatMock模式：使用ChatMock配置的模型名稱
                model = self.config.get('chatmock', 'model', fallback='gpt-5')
                return f"chat-{model}"  # 例如：chat-gpt-5
            else:
                # OpenAI模式：使用OpenAI配置的模型名稱
                model = self.config.get('openai', 'model', fallback='gpt-4')
                return model  # 例如：gpt-5-nano
            
        except Exception as e:
            logger.error(f"獲取模型顯示名稱失敗: {e}")
            # 返回默認值
            return "gpt-4"
    
    def load_prompt_template(self) -> str:
        """載入prompt模板"""
        prompt_file = 'prompt_template.txt'
        
        if not os.path.exists(prompt_file):
            logger.warning(f"Prompt文件不存在: {prompt_file}")
            return self.get_default_prompt()
        
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"載入prompt模板失敗: {e}")
            return self.get_default_prompt()
    
    def get_default_prompt(self) -> str:
        """獲取默認prompt模板"""
        return """你是一個佛學專家，專門負責對佛學問答進行精選評分。

請根據以下評分標準，對給定的問答內容進行評分：

**問題：** {title}
**回答：** {answer}

請按以下格式回答：

✅ **廣度評分：** XX分
✅ **深度評分：** XX分  
✅ **綜合評分：** XX分
✅ **廣度評論：** (100字以內)
✅ **深度評論：** (100字以內)
✅ **總體評價：** (80字以內)
✅ **問題摘要：** (50字以內)
✅ **回答摘要：** (100字以內)"""

    def load_excel_data(self) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
        """載入Excel數據"""
        file_path = self.config.get('excel', 'file_path')
        sheet_name = self.config.get('excel', 'sheet_name')
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook[sheet_name]
            logger.info(f"成功載入Excel文件: {file_path}, 工作表: {sheet_name}")
            return workbook, worksheet
        except Exception as e:
            logger.error(f"載入Excel文件失敗: {e}")
            raise

    def extract_qa_content(self, worksheet, row: int) -> Tuple[str, str]:
        """提取問答內容"""
        # 根據Excel結構調整列號
        question_col = self.config.getint('excel', 'question_column')
        answer_col = self.config.getint('excel', 'answer_column')
        
        try:
            question = worksheet.cell(row=row, column=question_col).value or ""
            answer = worksheet.cell(row=row, column=answer_col).value or ""
            
            return str(question).strip(), str(answer).strip()
        except Exception as e:
            logger.error(f"提取第 {row} 行內容失敗: {e}")
            return "", ""

    def evaluate_qa_quality(self, question: str, answer: str) -> Dict[str, Any]:
        """評估問答質量"""
        try:
            # 記錄開始時間
            start_time = time.time()
            logger.info(f"🤖 開始AI評分，問題長度: {len(question)}字，答案長度: {len(answer)}字")
            
            # 格式化提示詞
            prompt_start = time.time()
            formatted_prompt = self.prompt_template.format(title=question, answer=answer)
            prompt_time = time.time() - prompt_start
            logger.info(f"📝 提示詞格式化完成，耗時: {prompt_time:.2f}秒")
            
            # 準備API參數
            api_params = {
                'model': self.model,
                'messages': [{'role': 'user', 'content': formatted_prompt}],
                'temperature': self.temperature,
                'max_tokens': self.max_tokens
            }
            
            if self.max_tokens:
                api_params['max_tokens'] = self.max_tokens
            
            logger.info(f"🔧 API參數準備完成: 模型={self.model}, 溫度={self.temperature}")
            
            # 執行API調用
            logger.info(f"🌐 開始API調用...")
            api_start = time.time()
            
            # 添加重試機制
            max_retries = 3
            retry_count = 0
            last_error = None
            
            while retry_count < max_retries:
                try:
                    if retry_count > 0:
                        logger.info(f"🔄 第 {retry_count} 次重試...")
                        time.sleep(2 ** retry_count)  # 指數退避
                    
                    response = self.client.chat.completions.create(**api_params)
                    api_time = time.time() - api_start
                    logger.info(f"✅ API調用成功，耗時: {api_time:.2f}秒")
                    
                    # 檢查響應
                    if not response.choices or not response.choices[0].message:
                        raise ValueError("API響應格式異常")
                    
                    content = response.choices[0].message.content
                    logger.info(f"📄 收到AI響應，長度: {len(content)}字符")
                    
                    # 解析結果
                    logger.info(f"🔍 開始解析LLM評分結果...")
                    parse_start = time.time()
                    parsed_result = self.parse_evaluation_result(content)
                    parse_time = time.time() - parse_start
                    
                    # 統計解析結果
                    success_fields = sum(1 for v in parsed_result.values() if v != '解析失敗')
                    total_fields = len(parsed_result)
                    logger.info(f"✅ 解析完成: {success_fields}/{total_fields} 個字段成功，耗時: {parse_time:.2f}秒")
                    
                    # 計算總耗時
                    total_time = time.time() - start_time
                    logger.info(f"🎯 評分完成，總耗時: {total_time:.2f}秒")
                    
                    return parsed_result
                    
                except Exception as e:
                    last_error = e
                    retry_count += 1
                    api_time = time.time() - api_start
                    
                    if retry_count < max_retries:
                        logger.warning(f"⚠️ API調用失敗 (第{retry_count}次): {e}")
                        logger.warning(f"⏱️ 已耗時: {api_time:.2f}秒，準備重試...")
                    else:
                        logger.error(f"❌ API調用最終失敗，已重試{max_retries}次: {e}")
                        logger.error(f"⏱️ 總耗時: {api_time:.2f}秒")
                        break
            
            # 所有重試都失敗了
            logger.error(f"💥 AI評分完全失敗，返回錯誤結果")
            return {
                'breadth_score': 'API調用失敗',
                'depth_score': 'API調用失敗',
                'uniqueness_score': 'API調用失敗',
                'overall_score': 'API調用失敗',
                'breadth_comment': f'API調用失敗: {str(last_error)}',
                'depth_comment': f'API調用失敗: {str(last_error)}',
                'uniqueness_comment': f'API調用失敗: {str(last_error)}',
                'overall_comment': f'API調用失敗: {str(last_error)}',
                'question_summary': 'API調用失敗',
                'answer_summary': 'API調用失敗',
                'status': 'error'
            }
            
        except Exception as e:
            logger.error(f"❌ 評分過程發生未預期錯誤: {e}")
            return {
                'breadth_score': '系統錯誤',
                'depth_score': '系統錯誤',
                'uniqueness_score': '系統錯誤',
                'overall_score': '系統錯誤',
                'breadth_comment': f'系統錯誤: {str(e)}',
                'depth_comment': f'系統錯誤: {str(e)}',
                'uniqueness_comment': f'系統錯誤: {str(e)}',
                'overall_comment': f'系統錯誤: {str(e)}',
                'question_summary': '系統錯誤',
                'answer_summary': '系統錯誤',
                'status': 'error'
            }

    def parse_evaluation_result(self, result_text: str) -> Dict:
        """解析LLM的評分結果"""
        try:
            logger.info("開始解析LLM評分結果...")
            logger.debug(f"原始結果文本長度: {len(result_text)}")
            logger.debug(f"原始結果文本前500字符: {result_text[:500]}")
            
            # 初始化結果字典
            parsed_result = {
                'breadth_score': '解析失敗',
                'depth_score': '解析失敗', 
                'uniqueness_score': '解析失敗',
                'overall_score': '解析失敗',
                'breadth_comment': '解析失敗',
                'depth_comment': '解析失敗',
                'uniqueness_comment': '解析失敗',
                'overall_comment': '解析失敗',
                'question_summary': '解析失敗',
                'answer_summary': '解析失敗',
                'status': 'success'  # 添加狀態字段
            }
            
            # 改進的正則表達式，匹配LLM的實際輸出格式（支持多種格式）
            patterns = {
                'breadth_score': [
                    r'✅ \*\*廣度評分：\*\* (\d+)分',  # 繁體中文，有**標記
                    r'✅ 廣度評分：(\d+)分',           # 繁體中文，無**標記
                    r'✅ \*\*广度评分：\*\* (\d+)分',  # 簡體中文，有**標記
                    r'✅ 广度评分：(\d+)分'            # 簡體中文，無**標記
                ],
                'depth_score': [
                    r'✅ \*\*深度評分：\*\* (\d+)分',  # 繁體中文，有**標記
                    r'✅ 深度評分：(\d+)分',           # 繁體中文，無**標記
                    r'✅ \*\*深度评分：\*\* (\d+)分',  # 簡體中文，有**標記
                    r'✅ 深度评分：(\d+)分'            # 簡體中文，無**標記
                ],
                'uniqueness_score': [
                    r'✅ \*\*獨特性評分：\*\* (\d+)分',  # 繁體中文，有**標記
                    r'✅ 獨特性評分：(\d+)分',           # 繁體中文，無**標記
                    r'✅ \*\*独特性评分：\*\* (\d+)分',  # 簡體中文，有**標記
                    r'✅ 独特性评分：(\d+)分'            # 簡體中文，無**標記
                ],
                'overall_score': [
                    r'✅ \*\*綜合評分：\*\* (\d+)分',  # 繁體中文，有**標記
                    r'✅ 綜合評分：(\d+)分',           # 繁體中文，無**標記
                    r'✅ \*\*综合评分：\*\* (\d+)分',  # 簡體中文，有**標記
                    r'✅ 综合评分：(\d+)分'            # 簡體中文，無**標記
                ],
                'breadth_comment': [
                    r'✅ \*\*廣度評論：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 廣度評論：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*广度评论：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 广度评论：\s*\n(.+?)(?=\n\n|✅|$)'
                ],
                'depth_comment': [
                    r'✅ \*\*深度評論：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 深度評論：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*深度评论：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 深度评论：\s*\n(.+?)(?=\n\n|✅|$)'
                ],
                'uniqueness_comment': [
                    r'✅ \*\*獨特性評論：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 獨特性評論：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*独特性评论：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 独特性评论：\s*\n(.+?)(?=\n\n|✅|$)'
                ],
                'overall_comment': [
                    r'✅ \*\*總體評價：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 總體評價：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*总体评价：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 总体评价：\s*\n(.+?)(?=\n\n|✅|$)'
                ],
                'question_summary': [
                    r'✅ \*\*問題摘要：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 問題摘要：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*问题摘要：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 问题摘要：\s*\n(.+?)(?=\n\n|✅|$)'
                ],
                'answer_summary': [
                    r'✅ \*\*回答摘要：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 回答摘要：\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ \*\*回答摘要：\*\*\s*\n(.+?)(?=\n\n|✅|$)',
                    r'✅ 回答摘要：\s*\n(.+?)(?=\n\n|✅|$)'
                ]
            }
            
            # 嘗試解析每個字段
            for field, pattern_list in patterns.items():
                found_match = False
                for pattern in pattern_list:
                    try:
                        match = re.search(pattern, result_text, re.DOTALL | re.MULTILINE)
                        if match:
                            if 'score' in field:
                                # 分數字段
                                parsed_result[field] = int(match.group(1))
                                logger.debug(f"成功解析 {field}: {parsed_result[field]}")
                            else:
                                # 評論和摘要字段
                                parsed_result[field] = match.group(1).strip()
                                logger.debug(f"成功解析 {field}: {parsed_result[field][:50]}...")
                            found_match = True
                            break # 找到匹配後立即退出內層循環
                    except Exception as e:
                        logger.warning(f"嘗試模式 '{pattern}' 解析 {field} 失敗: {e}")
                        continue
                
                if not found_match:
                    logger.warning(f"未找到 {field} 的匹配")
                    # 嘗試更寬鬆的匹配
                    if 'score' in field:
                        # 嘗試其他可能的格式
                        alt_patterns = [
                            rf'{field.replace("_", "")}.*?(\d+)',
                            rf'{field.replace("_", "")}.*?(\d+)',
                            rf'(\d+).*?{field.replace("_", "")}'
                        ]
                        for alt_pattern in alt_patterns:
                            alt_match = re.search(alt_pattern, result_text, re.IGNORECASE)
                            if alt_match:
                                parsed_result[field] = int(alt_match.group(1))
                                logger.info(f"使用備用模式成功解析 {field}: {parsed_result[field]}")
                                break
                    else:
                        # 嘗試更寬鬆的文本匹配
                        alt_patterns = [
                            rf'{field.replace("_", "")}.*?([^\n]+)',
                            rf'([^\n]+).*?{field.replace("_", "")}'
                        ]
                        for alt_pattern in alt_patterns:
                            alt_match = re.search(alt_pattern, result_text, re.IGNORECASE)
                            if alt_match:
                                parsed_result[field] = alt_match.group(1).strip()
                                logger.info(f"使用備用模式成功解析 {field}: {parsed_result[field][:50]}...")
                                break
            
            # 計算綜合評分（加權平均）
            try:
                breadth = int(parsed_result['breadth_score'])
                depth = int(parsed_result['depth_score'])
                uniqueness = int(parsed_result['uniqueness_score'])
                
                # 加權平均：廣度30%，深度40%，獨特性30%
                overall_score = breadth * 0.3 + depth * 0.4 + uniqueness * 0.3
                parsed_result['overall_score'] = round(overall_score)
                logger.info(f"✅ 綜合評分計算完成: {breadth}×0.3 + {depth}×0.4 + {uniqueness}×0.3 = {overall_score:.1f} → {parsed_result['overall_score']}")
                
            except (ValueError, TypeError) as e:
                logger.warning(f"⚠️ 綜合評分計算失敗: {e}")
                parsed_result['overall_score'] = '計算失敗'
            
            # 檢查解析結果
            success_count = sum(1 for v in parsed_result.values() if v != '解析失敗')
            total_count = len(parsed_result)
            logger.info(f"解析完成: {success_count}/{total_count} 個字段成功")
            
            if success_count == 0:
                logger.error("所有字段解析失敗，請檢查LLM輸出格式")
                logger.error(f"完整結果文本: {result_text}")
            
            return parsed_result
            
        except Exception as e:
            logger.error(f"解析評分結果失敗: {e}")
            return {
                'breadth_score': '解析失敗',
                'depth_score': '解析失敗',
                'uniqueness_score': '解析失敗',
                'overall_score': '解析失敗', 
                'breadth_comment': '解析失敗',
                'depth_comment': '解析失敗',
                'uniqueness_comment': '解析失敗',
                'overall_comment': '解析失敗',
                'question_summary': '解析失敗',
                'answer_summary': '解析失敗'
            }

    def load_existing_results(self, results_file: str) -> Dict:
        """載入已有的精選評分結果"""
        if not os.path.exists(results_file):
            return {}
        
        try:
            with open(results_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('results', {})
        except Exception as e:
            logger.error(f"載入已有結果失敗: {e}")
            return {}

    def save_results(self, results_file: str):
        """保存精選評分結果到JSON文件"""
        try:
            self.processing_metadata['processing_end_time'] = datetime.now().isoformat()
            self.processing_metadata['total_processed'] = len(self.curation_results)
            self.processing_metadata['total_success'] = sum(1 for r in self.curation_results.values() if r.get('status') == 'success')
            self.processing_metadata['total_failed'] = len(self.curation_results) - self.processing_metadata['total_success']
            
            output_data = {
                'metadata': self.processing_metadata,
                'results': self.curation_results
            }
            
            with open(results_file, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"結果已保存到: {results_file}")
            logger.info(f"總處理: {self.processing_metadata['total_processed']}, 成功: {self.processing_metadata['total_success']}, 失敗: {self.processing_metadata['total_failed']}")
            
        except Exception as e:
            logger.error(f"保存結果失敗: {e}")

    def _get_filter_conditions(self) -> Dict:
        """獲取過濾條件"""
        try:
            conditions = {}
            
            # 檢查是否有特定的過濾條件
            if self.config.has_section('filter'):
                # 列值過濾條件（基於Excel列F、G、H的值）
                if self.config.has_option('filter', 'column_f_value'):
                    column_f_value = self.config.get('filter', 'column_f_value')
                    if column_f_value:
                        conditions['column_f_value'] = column_f_value.strip()
                
                if self.config.has_option('filter', 'column_g_value'):
                    column_g_value = self.config.get('filter', 'column_g_value')
                    if column_g_value:
                        conditions['column_g_value'] = column_g_value.strip()
                
                if self.config.has_option('filter', 'column_h_value'):
                    column_h_value = self.config.get('filter', 'column_h_value')
                    if column_h_value:
                        conditions['column_h_value'] = column_h_value.strip()
                
                # 檢查是否至少設置了一個列值過濾條件
                if not any(key in conditions for key in ['column_f_value', 'column_g_value', 'column_h_value']):
                    logger.warning("過濾模式下未設置任何列值過濾條件，建議設置至少一個列值")
            
            return conditions
            
        except Exception as e:
            logger.error(f"獲取過濾條件失敗: {e}")
            return {}

    def get_filtered_rows(self, worksheet) -> List[int]:
        """獲取過濾後的行號列表"""
        try:
            # 檢查是否啟用過濾模式
            use_filter_mode = self.config.getboolean('processing', 'use_filter_mode', fallback=False)
            if not use_filter_mode:
                return []
            
            logger.info("開始執行過濾模式...")
            
            # 獲取過濾條件
            filter_conditions = self._get_filter_conditions()
            logger.info(f"過濾條件: {filter_conditions}")
            
            # 檢查是否有列值過濾條件
            has_column_filters = any(key in filter_conditions for key in ['column_f_value', 'column_g_value', 'column_h_value'])
            
            if has_column_filters:
                # 使用快速過濾模式
                logger.info("使用快速列值過濾模式...")
                filtered_rows = self._fast_column_filter(worksheet, filter_conditions)
            else:
                # 使用傳統掃描模式
                logger.info("使用傳統掃描模式...")
                filtered_rows = self._traditional_scan_filter(worksheet)
            
            logger.info(f"過濾完成，共找到 {len(filtered_rows)} 行")
            
            # 記錄過濾結果的詳細信息
            if filtered_rows:
                logger.info(f"過濾結果行號: {filtered_rows[:10]}{'...' if len(filtered_rows) > 10 else ''}")
            
            return filtered_rows
            
        except Exception as e:
            logger.error(f"獲取過濾行失敗: {e}")
            return []

    def _fast_column_filter(self, worksheet, conditions: Dict) -> List[int]:
        """快速列值過濾模式 - 從Column H開始判斷以提高效率"""
        try:
            logger.info("開始快速列值過濾（從Column H開始）...")
            
            # 檢查緩存
            if self.filter_cache:
                excel_file = self.config.get('excel', 'file_path')
                f_value = conditions.get('column_f_value', '')
                g_value = conditions.get('column_g_value', '')
                h_value = conditions.get('column_h_value', '')
                
                cached_rows = self.filter_cache.get_cached_result(excel_file, f_value, g_value, h_value)
                if cached_rows:
                    logger.info(f"緩存命中！直接返回 {len(cached_rows)} 行過濾結果")
                    return cached_rows
                else:
                    logger.info("緩存未命中，開始掃描Excel文件")
            
            # 記錄使用的過濾條件
            used_conditions = []
            if 'column_f_value' in conditions:
                used_conditions.append(f"F列={conditions['column_f_value']}")
            if 'column_g_value' in conditions:
                used_conditions.append(f"G列={conditions['column_g_value']}")
            if 'column_h_value' in conditions:
                used_conditions.append(f"H列={conditions['column_h_value']}")
            
            logger.info(f"使用的列值過濾條件: {', '.join(used_conditions)}")
            
            # 獲取評分範圍設定
            start_index = self.config.getint('filter', 'start_index', fallback=0)
            end_index = self.config.getint('filter', 'end_index', fallback=0)
            score_all_filtered = self.config.getboolean('filter', 'score_all_filtered', fallback=False)
            
            # 計算需要的過濾條目數量
            if score_all_filtered:
                # 全部評分模式
                required_count = float('inf')  # 無限大，表示需要所有結果
                logger.info(f"評分設定: 全部評分模式，將評分所有過濾結果")
            elif end_index == 0:
                # 只評分第一條
                required_count = 1
                logger.info(f"評分設定: 只評分第一條過濾結果")
            else:
                # 評分指定範圍
                required_count = end_index - start_index + 1
                logger.info(f"評分設定: 評分第{start_index}到第{end_index}條過濾結果，共需{required_count}條")
            
            # 直接讀取列F、G、H的值進行過濾
            max_row = worksheet.max_row
            logger.info(f"Excel總行數: {max_row}")
            
            # 從第7行開始掃描（跳過標題行和說明行）
            scan_start = 7
            
            # 根據配置決定掃描範圍
            scan_full_file = self.config.getboolean('filter', 'scan_full_file', fallback=True)
            if scan_full_file:
                scan_end = max_row  # 掃描完整文件以建立完整緩存
                logger.info("🔍 掃描策略: 完整文件掃描（建立完整緩存）")
            else:
                scan_end = min(max_row, 1000)  # 限制掃描範圍以控制性能
                logger.info("⚠️ 掃描策略: 限制掃描範圍（緩存不完整，不推薦）")
            
            logger.info(f"掃描範圍: 第{scan_start}行到第{scan_end}行")
            logger.info(f"預計掃描行數: {scan_end - scan_start + 1}")
            
            # 計算預期的進度更新點（根據文件大小動態調整）
            expected_progress_points = []
            if scan_end - scan_start > 1000:
                # 大文件：每500行更新一次
                step = 500
            elif scan_end - scan_start > 500:
                # 中等文件：每200行更新一次
                step = 200
            else:
                # 小文件：每100行更新一次
                step = 100
            
            for i in range(step, scan_end + 1, step):
                if i >= scan_start:
                    expected_progress_points.append(i)
            
            logger.info(f"進度更新頻率: 每{step}行，預期進度更新點: {expected_progress_points[:10]}{'...' if len(expected_progress_points) > 10 else ''}")
            
            # 記錄開始時間
            import time
            start_time = time.time()
            last_progress_time = start_time
            
            filtered_rows = []
            
            for row in range(scan_start, scan_end + 1):
                try:
                    # 優化策略：從Column H開始判斷，因為H是最細分的第三級目錄
                    # 如果H不匹配，很可能F和G也不匹配，可以跳過後續檢查
                    matches = True
                    
                    # 1. 首先檢查第H列（第8列）- 第三級目錄
                    if 'column_h_value' in conditions:
                        cell_value = worksheet.cell(row=row, column=8).value
                        if cell_value is None:
                            cell_value = ""
                        if str(cell_value).strip() != conditions['column_h_value']:
                            matches = False
                            # H列不匹配，跳過後續檢查
                            continue
                        else:
                            logger.debug(f"第{row}行H列匹配: {cell_value}")
                    
                    # 2. 如果H列匹配，檢查第G列（第7列）- 第二級目錄
                    if matches and 'column_g_value' in conditions:
                        cell_value = worksheet.cell(row=row, column=7).value
                        if cell_value is None:
                            cell_value = ""
                        if str(cell_value).strip() != conditions['column_g_value']:
                            matches = False
                            # G列不匹配，跳過F列檢查
                            continue
                        else:
                            logger.debug(f"第{row}行G列匹配: {cell_value}")
                    
                    # 3. 如果G列也匹配，檢查第F列（第6列）- 第一級目錄
                    if matches and 'column_f_value' in conditions:
                        cell_value = worksheet.cell(row=row, column=6).value
                        if cell_value is None:
                            cell_value = ""
                        if str(cell_value).strip() != conditions['column_f_value']:
                            matches = False
                            # F列不匹配，該行不符合條件
                            continue
                        else:
                            logger.debug(f"第{row}行F列匹配: {cell_value}")
                    
                    # 所有設置的列值條件都匹配
                    if matches:
                        filtered_rows.append(row)
                        logger.debug(f"第{row}行通過所有列值過濾")
                        
                        # 檢查是否已達到目標數量（僅用於日誌，不停止掃描）
                        if len(filtered_rows) >= required_count and not score_all_filtered:
                            logger.info(f"已找到足夠的過濾結果: {len(filtered_rows)}條，目標: {required_count}條，繼續掃描以建立完整緩存")
                    
                    # 進度更新：根據文件大小動態調整更新頻率
                    if scan_end - scan_start > 1000:
                        # 大文件：每500行更新一次
                        progress_step = 500
                    elif scan_end - scan_start > 500:
                        # 中等文件：每200行更新一次
                        progress_step = 200
                    else:
                        # 小文件：每100行更新一次
                        progress_step = 100
                    
                    if row % progress_step == 0:
                        current_time = time.time()
                        elapsed_time = current_time - start_time
                        rows_per_second = row / elapsed_time if elapsed_time > 0 else 0
                        target_info = "全部" if required_count == float('inf') else f"目標 {required_count} 行（完整掃描建立緩存）"
                        logger.info(f"快速過濾進度: 已掃描到第 {row} 行，當前找到 {len(filtered_rows)} 行匹配，{target_info}，耗時 {elapsed_time:.1f}秒，速度 {rows_per_second:.1f}行/秒")
                        last_progress_time = current_time
                    
                    # 每100行也更新一次（作為主要進度點，適用於所有文件大小）
                    if row % 100 == 0:
                        current_time = time.time()
                        elapsed_time = current_time - start_time
                        rows_per_second = row / elapsed_time if elapsed_time > 0 else 0
                        target_info = "全部" if required_count == float('inf') else f"目標 {required_count} 行（完整掃描建立緩存）"
                        logger.info(f"快速過濾主要進度: 已掃描到第 {row} 行，當前找到 {len(filtered_rows)} 行匹配，{target_info}，耗時 {elapsed_time:.1f}秒，速度 {rows_per_second:.1f}行/秒")
                    
                    # 每500行更新一次（作為大進度點，適用於大文件）
                    if row % 500 == 0:
                        current_time = time.time()
                        elapsed_time = current_time - start_time
                        rows_per_second = row / elapsed_time if elapsed_time > 0 else 0
                        target_info = "全部" if required_count == float('inf') else f"目標 {required_count} 行（完整掃描建立緩存）"
                        logger.info(f"快速過濾大進度: 已掃描到第 {row} 行，當前找到 {len(filtered_rows)} 行匹配，{target_info}，耗時 {elapsed_time:.1f}秒，速度 {rows_per_second:.1f}行/秒")
                    
                    # 如果超過5秒沒有進度更新，強制輸出一次
                    current_time = time.time()
                    if current_time - last_progress_time > 5:
                        target_info = "全部" if required_count == float('inf') else f"目標 {required_count} 行（完整掃描建立緩存）"
                        logger.info(f"強制進度更新: 已掃描到第 {row} 行，當前找到 {len(filtered_rows)} 行匹配，{target_info}，耗時 {current_time - start_time:.1f}秒")
                        last_progress_time = current_time
                
                except Exception as e:
                    logger.warning(f"快速過濾第 {row} 行時出錯: {e}")
                    continue
            
            # 掃描完成後的總結日誌
            total_scanned = row - scan_start + 1 if 'row' in locals() else 0
            if score_all_filtered:
                logger.info(f"快速列值過濾完成，全部評分模式，找到 {len(filtered_rows)} 行匹配")
                logger.info(f"掃描統計: 從第{scan_start}行到第{scan_end}行，共掃描{scan_end - scan_start + 1}行")
            elif len(filtered_rows) >= required_count:
                logger.info(f"快速列值過濾完成，已找到足夠的結果: {len(filtered_rows)}條，目標: {required_count}條")
                logger.info(f"掃描統計: 從第{scan_start}行到第{scan_end}行，共掃描{scan_end - scan_start + 1}行（完整掃描以建立緩存）")
            else:
                logger.info(f"快速列值過濾完成，找到 {len(filtered_rows)} 行匹配，目標: {required_count}行")
                logger.info(f"掃描統計: 從第{scan_start}行到第{scan_end}行，共掃描{scan_end - scan_start + 1}行")
            
            # 計算過濾效率統計
            if score_all_filtered:
                # 全部評分模式，使用完整掃描範圍
                total_scanned = scan_end - scan_start + 1
            efficiency = (len(filtered_rows) / total_scanned) * 100 if total_scanned > 0 else 0
            logger.info(f"過濾效率: {efficiency:.2f}% ({len(filtered_rows)}/{total_scanned})")
            
            # 保存緩存結果
            if self.filter_cache and filtered_rows:
                excel_file = self.config.get('excel', 'file_path')
                f_value = conditions.get('column_f_value', '')
                g_value = conditions.get('column_g_value', '')
                h_value = conditions.get('column_h_value', '')
                
                # 準備掃描統計信息
                scan_stats = {
                    'scan_start': scan_start,
                    'scan_end': row if 'row' in locals() else scan_end,
                    'total_scanned': total_scanned,
                    'efficiency': efficiency,
                    'scan_time': time.time() - start_time if 'start_time' in locals() else 0
                }
                
                self.filter_cache.save_filter_result(excel_file, f_value, g_value, h_value, filtered_rows, scan_stats)
                logger.info(f"緩存已保存，共 {len(filtered_rows)} 行結果")
            
            return filtered_rows
            
        except Exception as e:
            logger.error(f"快速列值過濾失敗: {e}")
            return []

    def _traditional_scan_filter(self, worksheet) -> List[int]:
        """傳統掃描過濾模式"""
        try:
            logger.info("開始傳統掃描過濾...")
            
            # 獲取所有行
            logger.info("開始掃描Excel行...")
            all_rows = []
            
            # 優化：只掃描有數據的行，跳過空行
            max_row = worksheet.max_row
            logger.info(f"Excel總行數: {max_row}")
            
            # 從第7行開始掃描（跳過標題行和說明行）
            scan_start = 7
            
            # 根據配置決定掃描範圍
            scan_full_file = self.config.getboolean('filter', 'scan_full_file', fallback=True)
            if scan_full_file:
                scan_end = max_row  # 掃描完整文件以建立完整緩存
                logger.info("🔍 掃描策略: 完整文件掃描（建立完整緩存）")
            else:
                scan_end = min(max_row, 1000)  # 限制掃描範圍以控制性能
                logger.info("⚠️ 掃描策略: 限制掃描範圍（緩存不完整，不推薦）")
            
            logger.info(f"掃描範圍: 第{scan_start}行到第{scan_end}行")
            
            for row in range(scan_start, scan_end + 1):
                try:
                    # 快速檢查是否有內容（只檢查問題列）
                    question_col = self.config.getint('excel', 'question_column')
                    cell_value = worksheet.cell(row=row, column=question_col).value
                    
                    if cell_value and str(cell_value).strip():
                        all_rows.append(row)
                        
                        # 每100行記錄一次進度
                        if len(all_rows) % 100 == 0:
                            logger.info(f"已找到 {len(all_rows)} 行有內容的數據，當前掃描到第 {row} 行")
                    
                except Exception as e:
                    logger.warning(f"掃描第 {row} 行時出錯: {e}")
                    continue
            
            logger.info(f"掃描完成，找到 {len(all_rows)} 行有內容的數據")
            return all_rows
            
        except Exception as e:
            logger.error(f"傳統掃描過濾失敗: {e}")
            return []

    def process_batch(self, start_row: int = None, end_row: int = None, results_file: str = None):
        """批量處理問答精選評分，輸出到JSON文件"""
        # 記錄開始時間
        overall_start_time = time.time()
        logger.info(f"🚀 開始批量處理 - 時間: {datetime.now().strftime('%H:%M:%S')}")
        
        # 載入配置
        if start_row is None:
            start_row = self.config.getint('processing', 'start_row', fallback=2)
        if end_row is None:
            config_end_row = self.config.getint('processing', 'end_row', fallback=0)
            end_row = config_end_row if config_end_row > 0 else None
        
        if results_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            results_file = f'qa_curation_results_{timestamp}.json'
        
        logger.info(f"📁 結果文件: {results_file}")
        
        # 載入已有結果（支持續處理）
        logger.info("📂 載入已有結果...")
        load_start = time.time()
        self.curation_results = self.load_existing_results(results_file)
        load_time = time.time() - load_start
        logger.info(f"✅ 已有結果載入完成，耗時: {load_time:.2f}秒")
        
        # 載入Excel
        logger.info("📊 載入Excel數據...")
        excel_start = time.time()
        workbook, worksheet = self.load_excel_data()
        excel_time = time.time() - excel_start
        logger.info(f"✅ Excel數據載入完成，耗時: {excel_time:.2f}秒")
        
        # 檢查處理模式
        use_filter_mode = self.config.getboolean('processing', 'use_filter_mode', fallback=False)
        
        if use_filter_mode:
            # 過濾模式
            logger.info("🔍 使用過濾模式...")
            self.processing_metadata['processing_mode'] = "filter_mode"
            
            filter_start = time.time()
            rows_to_process = self.get_filtered_rows(worksheet)
            filter_time = time.time() - filter_start
            
            if not rows_to_process:
                logger.warning("⚠️ 過濾模式下沒有找到符合條件的行")
                return results_file
            
            logger.info(f"✅ 過濾完成，找到 {len(rows_to_process)} 行，耗時: {filter_time:.2f}秒")
            
            # 獲取過濾結果的評分範圍
            filter_start_index = self.config.getint('filter', 'start_index', fallback=0)
            filter_end_index = self.config.getint('filter', 'end_index', fallback=0)
            
            if filter_end_index > 0:
                # 指定範圍
                start_idx = max(0, filter_start_index)
                end_idx = min(len(rows_to_process), filter_end_index + 1)
                rows_to_process = rows_to_process[start_idx:end_idx]
                logger.info(f"🎯 過濾模式：處理第 {start_idx+1} 到第 {end_idx} 條過濾結果，共 {len(rows_to_process)} 條")
            else:
                # 只處理第一條
                rows_to_process = rows_to_process[:1]
                logger.info("🎯 過濾模式：只處理第一條過濾結果")
            
        else:
            # 傳統模式（指定行號）
            logger.info("📝 使用行號模式...")
            self.processing_metadata['processing_mode'] = "row_mode"
            
            # 確定處理範圍
            max_row = worksheet.max_row
            if end_row is None or end_row > max_row:
                end_row = max_row
            
            rows_to_process = list(range(start_row, end_row + 1))
            logger.info(f"🎯 行號模式：處理第 {start_row} 到 {end_row} 行，共 {len(rows_to_process)} 條記錄")
        
        # 開始評分處理
        total_count = len(rows_to_process)
        logger.info(f"🚀 開始評分處理，總目標: {total_count} 條記錄")
        
        # 顯示進度條
        self._display_progress_bar(0, total_count, "開始處理")
        
        processed_count = 0
        success_count = 0
        failed_count = 0
        skipped_count = 0
        
        # 記錄處理開始時間
        processing_start_time = time.time()
        last_save_time = processing_start_time
        
        for i, row in enumerate(rows_to_process):
            current_time = time.time()
            elapsed_time = current_time - processing_start_time
            
            # 計算進度和預估時間
            progress_percent = (i / total_count) * 100 if total_count > 0 else 0
            if i > 0:
                avg_time_per_item = elapsed_time / i
                remaining_items = total_count - i
                estimated_remaining_time = remaining_items * avg_time_per_item
                
                logger.info(f"📈 進度: {i+1}/{total_count} ({progress_percent:.1f}%) - 已耗時: {elapsed_time:.1f}秒")
                logger.info(f"⏳ 預估剩餘時間: {estimated_remaining_time:.1f}秒 ({estimated_remaining_time/60:.1f}分鐘)")
                logger.info(f"🚀 平均速度: {i/elapsed_time:.2f} 條/秒")
            
            # 更新進度條
            self._display_progress_bar(i + 1, total_count, f"處理第{i+1}條")
            
            try:
                # 檢查是否已處理
                row_key = str(row)
                if row_key in self.curation_results:
                    logger.info(f"⏭️ 第 {row} 行已處理，跳過")
                    skipped_count += 1
                    continue
                
                # 提取問答內容
                logger.info(f"📖 提取第 {row} 行問答內容...")
                extract_start = time.time()
                question, answer = self.extract_qa_content(worksheet, row)
                extract_time = time.time() - extract_start
                
                if not question and not answer:
                    logger.info(f"⚠️ 第 {row} 行無內容，跳過")
                    skipped_count += 1
                    continue
                
                logger.info(f"🔄 處理第 {row} 行: {question[:100]}...")
                logger.info(f"📊 內容提取耗時: {extract_time:.2f}秒")
                
                # 進行精選評分
                logger.info(f"🤖 開始AI評分...")
                scoring_start = time.time()
                result = self.evaluate_qa_quality(question, answer)
                scoring_time = time.time() - scoring_start
                logger.info(f"✅ AI評分完成，耗時: {scoring_time:.2f}秒")
                
                # 保存結果
                logger.info(f"💾 保存評分結果...")
                save_start = time.time()
                self.curation_results[row_key] = {
                    'row_number': row,
                    'question': question[:500],  # 限制長度
                    'answer': answer[:1000],     # 限制長度
                    'breadth_score': result.get('breadth_score', ''),
                    'depth_score': result.get('depth_score', ''),
                    'uniqueness_score': result.get('uniqueness_score', ''),
                    'overall_score': result.get('overall_score', ''),
                    'breadth_comment': result.get('breadth_comment', ''),
                    'depth_comment': result.get('depth_comment', ''),
                    'uniqueness_comment': result.get('uniqueness_comment', ''),
                    'overall_comment': result.get('overall_comment', ''),
                    'question_summary': result.get('question_summary', ''),
                    'answer_summary': result.get('answer_summary', ''),
                    'status': result.get('status', 'success'),  # 使用get方法，默認為success
                    'processed_time': datetime.now().isoformat()
                }
                save_time = time.time() - save_start
                logger.info(f"✅ 結果保存完成，耗時: {save_time:.2f}秒")
                
                processed_count += 1
                if result.get('status') == 'success':
                    success_count += 1
                
                # 計算總耗時
                total_item_time = extract_time + scoring_time + save_time
                logger.info(f"✅ 第 {row} 行處理完成，總耗時: {total_item_time:.2f}秒")
                
                # 每處理10條記錄保存一次
                if processed_count % 10 == 0:
                    logger.info(f"💾 執行中間保存...")
                    save_start = time.time()
                    self.save_results(results_file)
                    save_time = time.time() - save_start
                    last_save_time = time.time()
                    logger.info(f"✅ 中間保存完成，已處理 {processed_count} 條記錄，保存耗時: {save_time:.2f}秒")
                
                # API調用間隔
                if i < total_count - 1:  # 不是最後一條
                    logger.info(f"⏸️ 等待1秒後處理下一條...")
                    time.sleep(1)
                
            except Exception as e:
                logger.error(f"❌ 處理第 {row} 行時發生錯誤: {e}")
                failed_count += 1
                processed_count += 1
                continue
        
        # 最終保存
        logger.info(f"💾 執行最終保存...")
        final_save_start = time.time()
        self.save_results(results_file)
        final_save_time = time.time() - final_save_start
        
        # 計算總統計
        total_time = time.time() - overall_start_time
        processing_time = time.time() - processing_start_time
        
        logger.info(f"🎉 批量處理完成！")
        logger.info(f"📊 統計結果:")
        logger.info(f"   - 總計: {total_count} 條")
        logger.info(f"   - 成功: {success_count} 條")
        logger.info(f"   - 失敗: {failed_count} 條")
        logger.info(f"   - 跳過: {skipped_count} 條")
        logger.info(f"⏱️ 時間統計:")
        logger.info(f"   - 總耗時: {total_time:.2f}秒 ({total_time/60:.1f}分鐘)")
        logger.info(f"   - 處理耗時: {processing_time:.2f}秒 ({processing_time/60:.1f}分鐘)")
        logger.info(f"   - 最終保存耗時: {final_save_time:.2f}秒")
        if processed_count > 0:
            logger.info(f"🚀 性能統計:")
            logger.info(f"   - 平均速度: {processed_count/processing_time:.2f} 條/秒")
            logger.info(f"   - 平均每條耗時: {processing_time/processed_count:.2f} 秒")
        
        return results_file

    def _display_progress_bar(self, current: int, total: int, status: str = ""):
        """顯示進度條"""
        try:
            if total <= 0:
                return
            
            # 計算進度百分比
            progress = (current / total) * 100
            
            # 進度條長度
            bar_length = 30
            filled_length = int(bar_length * current // total)
            
            # 構建進度條
            bar = '█' * filled_length + '░' * (bar_length - filled_length)
            
            # 顯示進度條
            print(f"\r📊 進度: [{bar}] {current}/{total} ({progress:.1f}%) - {status}", end='', flush=True)
            
            # 如果完成，換行
            if current >= total:
                print()
                
        except Exception as e:
            logger.warning(f"進度條顯示失敗: {e}")

def main():
    """主函數"""
    parser = argparse.ArgumentParser(
        description="佛學問答精選自動化系統",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 使用OpenAI API（推薦）
  python3 qa_curator.py --api-key YOUR_API_KEY --api-type openai
  
  # 使用ChatMock
  python3 qa_curator.py --api-type chatmock
  
  # 使用環境變量
  export OPENAI_API_KEY=YOUR_API_KEY
  python3 qa_curator.py --api-type openai
  
  # 使用配置文件（不推薦，會產生commit警告）
  python3 qa_curator.py
        """
    )
    
    parser.add_argument(
        '--api-key',
        type=str,
        help='OpenAI API Key（推薦使用此方式）'
    )
    
    parser.add_argument(
        '--api-type',
        type=str,
        choices=['openai', 'chatmock'],
        help='API類型選擇：openai 或 chatmock（覆蓋配置文件設置）'
    )
    
    parser.add_argument(
        '--chatmock-url',
        type=str,
        help='ChatMock服務器地址（覆蓋配置文件設置）'
    )
    
    parser.add_argument(
        '--config',
        type=str,
        default='config.ini',
        help='配置文件路徑（默認: config.ini）'
    )
    
    args = parser.parse_args()
    
    print("佛學問答精選自動化系統")
    print("=" * 50)
    
    try:
        curator = BuddhistQACurator(
            config_file=args.config,
            api_key=args.api_key,
            api_type=args.api_type,
            chatmock_url=args.chatmock_url
        )
        
        # 處理指定範圍
        results_file = curator.process_batch()
        
        print(f"\n✅ 精選評分完成！結果已保存到: {results_file}")
        print("接下來請運行 results_to_excel.py 將結果寫入Excel文件")
        
    except Exception as e:
        logger.error(f"程序執行失敗: {e}")
        print(f"❌ 程序執行失敗: {e}")

if __name__ == "__main__":
    main()
