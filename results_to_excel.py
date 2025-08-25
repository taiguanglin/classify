#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分類結果寫入Excel程序
讀取JSON格式的分類結果，批量寫入Excel文件
"""

import json
import openpyxl
from openpyxl import load_workbook
import configparser
import logging
from datetime import datetime
import os
import argparse
from typing import Dict, Any
from copy import copy
try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("警告: tqdm库未安装，将使用简单进度显示。建议安装: pip install tqdm")

# 設置日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ResultsToExcel:
    """分類結果寫入Excel"""
    
    def __init__(self, config_file: str = 'config.ini'):
        """初始化"""
        self.config = configparser.ConfigParser()
        self.config.read(config_file, encoding='utf-8')
        
        logger.info("Excel寫入器初始化完成")
    
    def load_results(self, results_file: str) -> Dict[str, Any]:
        """載入分類結果"""
        if not os.path.exists(results_file):
            raise FileNotFoundError(f"結果文件不存在: {results_file}")
        
        try:
            with open(results_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            logger.info(f"成功載入結果文件: {results_file}")
            logger.info(f"元數據: 總處理 {data['metadata'].get('total_processed', 0)}, "
                       f"成功 {data['metadata'].get('total_success', 0)}")
            
            return data
        except Exception as e:
            logger.error(f"載入結果文件失敗: {e}")
            raise
    
    def create_output_excel(self, source_file: str, output_file: str, required_rows: set = None) -> tuple:
        """創建輸出Excel文件，只包含需要的行"""
        try:
            print("📁 正在載入Excel文件...")
            
            # 載入原始文件
            source_workbook = load_workbook(source_file)
            sheet_name = self.config.get('excel', 'sheet_name')
            source_worksheet = source_workbook[sheet_name]
            
            print("✅ Excel文件載入完成")
            print("🧹 正在創建精簡工作表...")
            
            # 創建新的工作簿，只包含需要的行
            workbook, worksheet = self._create_minimal_excel(source_worksheet, required_rows)
            
            # 清理工作表，只保留指定的工作表
            self._clean_worksheets(workbook, sheet_name)
            
            logger.info(f"成功創建精簡Excel文件，包含 {len(required_rows) + 1 if required_rows else 0} 行")
            return workbook, worksheet
        except Exception as e:
            logger.error(f"創建輸出Excel失敗: {e}")
            raise
    
    def _create_minimal_excel(self, source_worksheet, required_rows: set):
        """創建只包含需要行的新Excel工作簿"""
        from openpyxl import Workbook
        
        # 創建新工作簿
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = source_worksheet.title
        
        # 總是包含標題行(第6行)
        rows_to_copy = {6}
        if required_rows:
            rows_to_copy.update(required_rows)
        
        # 記錄標題行的新位置
        self.title_row_new = None
        
        # 獲取源工作表的最大列數
        max_col = source_worksheet.max_column
        
        # 創建行號映射（新行號 -> 原行號）
        self.row_mapping = {}
        new_row = 1
        
        # 按順序復制行
        for original_row in sorted(rows_to_copy):
            try:
                # 復制整行數據
                for col in range(1, max_col + 1):
                    source_cell = source_worksheet.cell(row=original_row, column=col)
                    target_cell = worksheet.cell(row=new_row, column=col)
                    
                    # 復制值
                    target_cell.value = source_cell.value
                    
                    # 復制格式（如果有的話）
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
                
                # 記錄行號映射
                self.row_mapping[original_row] = new_row
                
                # 記錄標題行的新位置
                if original_row == 6:
                    self.title_row_new = new_row
                
                new_row += 1
                
            except Exception as e:
                logger.warning(f"復制第 {original_row} 行時出錯: {e}")
        
        logger.info(f"成功創建精簡工作表，從 {len(rows_to_copy)} 行復制")
        return workbook, worksheet
    
    def _clean_worksheets(self, workbook, keep_sheet_name: str):
        """清理工作表，只保留指定的工作表"""
        try:
            sheets_to_remove = []
            for sheet_name in workbook.sheetnames:
                if sheet_name != keep_sheet_name:
                    sheets_to_remove.append(sheet_name)
            
            if sheets_to_remove:
                logger.info(f"將刪除 {len(sheets_to_remove)} 個工作表: {', '.join(sheets_to_remove)}")
                for sheet_name in sheets_to_remove:
                    del workbook[sheet_name]
                logger.info(f"只保留工作表: {keep_sheet_name}")
            else:
                logger.info(f"工作表 {keep_sheet_name} 已是最後一個工作表")
                
        except Exception as e:
            logger.error(f"清理工作表失敗: {e}")
            # 不拋出異常，讓程序繼續執行
    
    def _add_column_headers(self, worksheet):
        """在標題行添加新列的標題"""
        try:
            # 獲取輸出列配置
            classification_col = self.config.getint('output', 'classification_column')
            reason_col = self.config.getint('output', 'reason_column')
            directory1_col = self.config.getint('output', 'directory1_column')
            directory2_col = self.config.getint('output', 'directory2_column')
            directory3_col = self.config.getint('output', 'directory3_column')
            
            # 使用正確的標題行號
            title_row = getattr(self, 'title_row_new', 6)
            worksheet.cell(row=title_row, column=classification_col).value = "LLM分類"
            worksheet.cell(row=title_row, column=reason_col).value = "LLM分析原因"
            worksheet.cell(row=title_row, column=directory1_col).value = "第一層目錄"
            worksheet.cell(row=title_row, column=directory2_col).value = "第二層目錄"
            worksheet.cell(row=title_row, column=directory3_col).value = "第三層目錄"
            
            # 設置標題格式
            title_row = getattr(self, 'title_row_new', 6)
            for col in [classification_col, reason_col, directory1_col, directory2_col, directory3_col]:
                cell = worksheet.cell(row=title_row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal='center',
                    vertical='center'
                )
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            
            logger.info(f"已添加列標題: 第{classification_col}列(LLM分類), 第{reason_col}列(LLM分析原因), 第{directory1_col}列(第一層目錄), 第{directory2_col}列(第二層目錄), 第{directory3_col}列(第三層目錄)")
            
        except Exception as e:
            logger.error(f"添加列標題失敗: {e}")
            # 不拋出異常，讓程序繼續執行
    
    def _load_directory_system(self) -> Dict[str, Dict[str, str]]:
        """載入目錄體系"""
        try:
            directory_system = {}
            
            # 讀取prompt_template.txt文件
            prompt_file = 'prompt_template.txt'
            if not os.path.exists(prompt_file):
                logger.warning(f"Prompt文件不存在: {prompt_file}")
                return {}
            
            with open(prompt_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 解析目錄體系
            import re
            
            # 匹配第一層目錄：I. 【义理】、II. 【修心】等
            level1_pattern = r'^([IV]+)\.\s*【([^】]+)】'
            
            # 匹配第二層目錄：I.I. 真相、II.II. 打坐等
            level2_pattern = r'^\s*([IV]+\.[IV]+)\.\s*([^：]+)'
            
            lines = content.split('\n')
            current_level1 = None
            
            for line in lines:
                line = line.strip()
                
                # 檢查第一層目錄
                level1_match = re.match(level1_pattern, line)
                if level1_match:
                    roman_num = level1_match.group(1)
                    chinese_name = level1_match.group(2)
                    current_level1 = roman_num
                    
                    directory_system[roman_num] = {
                        'name': chinese_name,
                        'full_name': f"{roman_num}.【{chinese_name}】",
                        'subcategories': {}
                    }
                
                # 檢查第二層目錄
                elif current_level1:
                    level2_match = re.match(level2_pattern, line)
                    if level2_match:
                        roman_sub = level2_match.group(1)
                        chinese_sub = level2_match.group(2)
                        
                        if current_level1 in directory_system:
                            directory_system[current_level1]['subcategories'][roman_sub] = {
                                'name': chinese_sub,
                                'full_name': f"{roman_sub}. {chinese_sub}"
                            }
            
            logger.info(f"成功載入目錄體系，包含 {len(directory_system)} 個第一層分類")
            return directory_system
            
        except Exception as e:
            logger.error(f"載入目錄體系失敗: {e}")
            return {}
    
    def _parse_directory_from_classification(self, classification_text: str) -> Dict[str, str]:
        """從LLM分類結果中解析目錄信息"""
        try:
            directory_info = {
                'directory1': '',
                'directory2': '',
                'directory3': ''
            }
            
            if not classification_text:
                return directory_info
            
            # 分割分類結果（按換行符分割）
            lines = classification_text.strip().split('\n')
            
            # 取第一個分類（信心度最高的）
            if lines:
                first_classification = lines[0].strip()
                
                # 解析目錄結構
                # 例如："I.I. 真相（80%）" 或 "II.II. 打坐（信心度95%）"
                if '.' in first_classification:
                    # 有子分類的情況
                    # 使用正則表達式提取羅馬數字和分類名稱
                    import re
                    
                    # 匹配模式：I.I. 真相（80%）或 II.II. 打坐（信心度95%）
                    pattern = r'^([IV]+)\.([IV]+)\.\s*([^（]+)'
                    match = re.match(pattern, first_classification)
                    
                    if match:
                        level1_roman = match.group(1)
                        level2_roman = f"{match.group(1)}.{match.group(2)}"
                        level2_name = match.group(3).strip()
                        
                        # 使用目錄體系獲取完整名稱
                        if level1_roman in self.directory_system:
                            # 第一層：IV.【生活】
                            directory_info['directory1'] = self.directory_system[level1_roman]['full_name']
                            
                            # 第二層：IV.II. 工作
                            if level2_roman in self.directory_system[level1_roman]['subcategories']:
                                directory_info['directory2'] = self.directory_system[level1_roman]['subcategories'][level2_roman]['full_name']
                            else:
                                # 如果目錄體系中沒有找到對應的第二層，使用解析出的名稱
                                directory_info['directory2'] = f"{level2_roman}. {level2_name}"
                            
                            # 第三層：目前不存在，預留為空，為未來擴展做準備
                            directory_info['directory3'] = ''
                
                elif '【' in first_classification and '】' in first_classification:
                    # 只有大分類的情況，如"【义理】（85%）"
                    # 提取分類名稱
                    start = first_classification.find('【') + 1
                    end = first_classification.find('】')
                    if start > 0 and end > start:
                        category_name = first_classification[start:end].strip()
                        # 查找對應的羅馬數字
                        for roman, info in self.directory_system.items():
                            if info['name'] == category_name:
                                directory_info['directory1'] = info['full_name']
                                break
            
            logger.debug(f"解析目錄信息: {directory_info}")
            return directory_info
            
        except Exception as e:
            logger.error(f"解析目錄信息失敗: {e}")
            return {'directory1': '', 'directory2': '', 'directory3': ''}
    
    def write_classification_result(self, worksheet, row: int, result: Dict[str, Any]):
        """寫入分類結果到指定行"""
        try:
            # 使用行號映射獲取新的行號
            if hasattr(self, 'row_mapping') and row in self.row_mapping:
                actual_row = self.row_mapping[row]
            else:
                actual_row = row
            
            # 獲取列配置
            classification_col = self.config.getint('output', 'classification_column')
            reason_col = self.config.getint('output', 'reason_column')
            question_col = self.config.getint('excel', 'question_column')
            answer_col = self.config.getint('excel', 'answer_column')
            directory1_col = self.config.getint('output', 'directory1_column')
            directory2_col = self.config.getint('output', 'directory2_column')
            directory3_col = self.config.getint('output', 'directory3_column')
            
            # 從LLM分類結果中解析目錄信息
            directory_info = self._parse_directory_from_classification(result.get('classification', ''))
            
            # 寫入分類結果到輸出列
            self._write_cell_with_format(worksheet, actual_row, classification_col, result.get('classification', ''))
            self._write_cell_with_format(worksheet, actual_row, reason_col, result.get('reason', ''))
            
            # 寫入解析出的目錄信息
            self._write_cell_with_format(worksheet, actual_row, directory1_col, directory_info.get('directory1', ''))
            self._write_cell_with_format(worksheet, actual_row, directory2_col, directory_info.get('directory2', ''))
            self._write_cell_with_format(worksheet, actual_row, directory3_col, directory_info.get('directory3', ''))
            
            # 設置問題和答案的comment
            question_summary = result.get('question_summary', '')
            answer_summary = result.get('answer_summary', '')
            
            # 只在有摘要時才添加註釋
            if question_summary and question_summary.strip():
                self._set_cell_comment(worksheet, actual_row, question_col, question_summary, '問題重點摘要')
            
            if answer_summary and answer_summary.strip() and answer_summary != "僅基於問題分類，無法提供回答摘要":
                self._set_cell_comment(worksheet, actual_row, answer_col, answer_summary, '回答重點摘要')
            
        except Exception as e:
            logger.error(f"寫入第 {row} 行結果失敗: {e}")
            raise
    
    def _write_cell_with_format(self, worksheet, row: int, col: int, value: str):
        """寫入單元格並設置自動換行格式"""
        try:
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
            
            # 設置自動換行
            cell.alignment = openpyxl.styles.Alignment(
                wrap_text=True,
                vertical='top',
                horizontal='left'
            )
            
            # 設置邊框樣式
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
        except Exception as e:
            logger.error(f"設置單元格格式失敗 (行{row}, 列{col}): {e}")
            raise
    
    def _set_cell_comment(self, worksheet, row: int, col: int, comment_text: str, author: str):
        """設置單元格comment"""
        try:
            if comment_text and comment_text.strip():
                cell = worksheet.cell(row=row, column=col)
                
                # 添加"LLM摘要:"前缀
                formatted_text = f"LLM摘要:\n{comment_text}"
                
                # 創建comment對象
                comment = openpyxl.comments.Comment(
                    text=formatted_text,
                    author=author
                )
                
                # 設置comment樣式
                comment.width = 300  # 設置comment寬度
                comment.height = 150  # 設置comment高度
                
                # 將comment添加到單元格
                cell.comment = comment
                
        except Exception as e:
            logger.error(f"設置comment失敗 (行{row}, 列{col}): {e}")
            # 不拋出異常，讓程序繼續執行
    
    def process_results(self, results_file: str, output_file: str = None):
        """處理分類結果並寫入Excel"""
        # 載入結果
        data = self.load_results(results_file)
        results = data.get('results', {})
        metadata = data.get('metadata', {})
        
        if not results:
            logger.warning("沒有找到分類結果")
            return
        
        # 確定輸出文件名
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"classified_results_{timestamp}.xlsx"
        
        # 創建輸出Excel
        # 優先使用元數據中的源文件，如果不存在則使用配置文件中的文件
        metadata_source_file = metadata.get('source_file')
        config_source_file = self.config.get('excel', 'file_path')
        
        # 檢查元數據中的源文件是否存在
        if metadata_source_file and os.path.exists(metadata_source_file):
            source_file = metadata_source_file
            logger.info(f"使用元數據中的源文件: {source_file}")
        else:
            source_file = config_source_file
            if metadata_source_file:
                logger.warning(f"元數據中的源文件不存在: {metadata_source_file}")
                logger.info(f"使用配置文件中的源文件: {source_file}")
            else:
                logger.info(f"使用配置文件中的源文件: {source_file}")
        
        # 獲取需要的行號
        required_rows = set(int(row_key) for row_key in results.keys())
        
        workbook, worksheet = self.create_output_excel(source_file, output_file, required_rows)
        
        # 載入目錄體系
        self.directory_system = self._load_directory_system()
        
        # 添加新列的標題
        self._add_column_headers(worksheet)
        
        total_items = len(results)
        logger.info(f"開始寫入 {total_items} 條分類結果")
        print(f"📊 開始處理 {total_items} 條分類結果...")
        
        # 統計信息
        success_count = 0
        failed_count = 0
        
        # 按行號排序處理
        sorted_results = sorted(results.items(), key=lambda x: int(x[0]))
        
        # 使用進度條
        if TQDM_AVAILABLE:
            pbar = tqdm(sorted_results, desc="寫入分類結果", unit="條")
        else:
            pbar = sorted_results
            print("進度: [", end="")
        
        for i, (row_key, result) in enumerate(pbar):
            try:
                row_number = int(row_key)
                
                # 跳過標題行（第6行），從第7行開始寫入數據
                if row_number == 6:
                    if not TQDM_AVAILABLE:
                        print("=", end="", flush=True)
                    continue
                
                # 寫入結果
                self.write_classification_result(worksheet, row_number, result)
                
                if result.get('status') == 'success':
                    success_count += 1
                else:
                    failed_count += 1
                
                # 更新進度條
                if not TQDM_AVAILABLE:
                    print("=", end="", flush=True)
                
                # 每處理10條記錄顯示進度
                if (success_count + failed_count) % 10 == 0:
                    current_progress = success_count + failed_count
                    if TQDM_AVAILABLE:
                        pbar.set_postfix({
                            '成功': success_count,
                            '失敗': failed_count,
                            '進度': f"{current_progress}/{total_items}"
                        })
                    else:
                        print(f"\n進度: {current_progress}/{total_items} (成功: {success_count}, 失敗: {failed_count})", end="")
                
            except Exception as e:
                logger.error(f"處理行 {row_key} 時發生錯誤: {e}")
                failed_count += 1
                continue
        
        if not TQDM_AVAILABLE:
            print("] 完成!")
        
        print(f"✅ 數據寫入完成: 成功 {success_count} 條，失敗 {failed_count} 條")
        
        # 自動調整列寬和行高
        print("📏 正在調整列寬...")
        self._auto_adjust_columns_and_rows(worksheet)
        
        # 由於已經只輸出需要的行，不再需要隱藏行
        print("👁️ 輸出文件已經只包含需要的行，無需隱藏行...")
        logger.info("輸出文件已經只包含需要的行，無需隱藏行")
        
        # 保存Excel文件
        print("💾 正在保存Excel文件...")
        try:
            workbook.save(output_file)
            print("✅ Excel文件保存完成!")
            logger.info(f"✅ Excel文件已保存: {output_file}")
            logger.info(f"📊 統計: 成功寫入 {success_count} 條，失敗 {failed_count} 條")
            
            # 顯示元數據信息
            if metadata:
                logger.info("📋 處理信息:")
                logger.info(f"   源文件: {metadata.get('source_file', 'N/A')}")
                logger.info(f"   處理時間: {metadata.get('processing_start_time', 'N/A')} - {metadata.get('processing_end_time', 'N/A')}")
                logger.info(f"   總處理: {metadata.get('total_processed', 0)}")
                logger.info(f"   成功率: {metadata.get('total_success', 0)}/{metadata.get('total_processed', 0)}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"保存Excel文件失敗: {e}")
            raise
    
    def _auto_adjust_columns_and_rows(self, worksheet):
        """自動調整列寬和行高以適應內容"""
        try:
            # 獲取輸出列配置
            classification_col = self.config.getint('output', 'classification_column')
            reason_col = self.config.getint('output', 'reason_column')
            
            # 定義所有需要調整的列及其寬度範圍
            columns_to_adjust = [
                # 列G（答疑日期）- 需要足夠寬度顯示日期
                {'col': 7, 'min_width': 20, 'max_width': 25, 'name': '答疑日期'},
                # 列H（问题）- 適中寬度，避免過寬
                {'col': 8, 'min_width': 30, 'max_width': 60, 'name': '问题'},
                # 列I（答案）- 適中寬度，避免過寬
                {'col': 9, 'min_width': 30, 'max_width': 60, 'name': '答案'},
                # LLM分類列
                {'col': classification_col, 'min_width': 15, 'max_width': 50, 'name': 'LLM分類'},
                # LLM分析原因列
                {'col': reason_col, 'min_width': 15, 'max_width': 50, 'name': 'LLM分析原因'},
                # 第一層目錄
                {'col': 14, 'min_width': 15, 'max_width': 30, 'name': '第一層目錄'},
                # 第二層目錄
                {'col': 15, 'min_width': 15, 'max_width': 30, 'name': '第二層目錄'},
                # 第三層目錄
                {'col': 16, 'min_width': 15, 'max_width': 30, 'name': '第三層目錄'},
            ]
            
            # 調整所有列的寬度
            for col_config in columns_to_adjust:
                self._adjust_column_width(
                    worksheet, 
                    col_config['col'], 
                    col_config['max_width'], 
                    col_config['min_width'],
                    col_config['name']
                )
            
            logger.info("列寬自動調整完成")
            
        except Exception as e:
            logger.error(f"自動調整列寬失敗: {e}")
            # 不拋出異常，讓程序繼續執行
    
    def _adjust_column_width(self, worksheet, col: int, max_width: int, min_width: int, col_name: str = None):
        """調整單列寬度"""
        try:
            # 計算該列的最大內容長度
            max_length = min_width
            total_rows = worksheet.max_row
            
            # 使用進度條處理大量行
            if TQDM_AVAILABLE and total_rows > 1000:
                row_range = tqdm(range(1, total_rows + 1), desc=f"調整{col_name or f'列{openpyxl.utils.get_column_letter(col)}'}", leave=False)
            else:
                row_range = range(1, total_rows + 1)
            
            for row in row_range:
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    # 計算文本長度（中文字符算2個字符寬度）
                    text_length = self._calculate_text_width(str(cell.value))
                    max_length = max(max_length, text_length)
            
            # 限制最大寬度
            adjusted_width = min(max_length + 2, max_width)  # +2 為邊距
            
            # 設置列寬
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
            
            logger.debug(f"列 {col_name or openpyxl.utils.get_column_letter(col)} 寬度調整為: {adjusted_width}")
            
        except Exception as e:
            logger.error(f"調整列 {col_name or col} 寬度失敗: {e}")
    
    def _calculate_text_width(self, text: str) -> int:
        """計算文本寬度（中文字符算2個字符寬度）"""
        width = 0
        for char in text:
            if ord(char) > 127:  # 中文字符
                width += 2
            else:  # 英文字符
                width += 1
        return width
    
    def _hide_rows_7_to_539(self, worksheet):
        """隱藏第7行到第539行"""
        try:
            # 隱藏從第7行到第539行
            for row_num in range(7, 540):
                worksheet.row_dimensions[row_num].hidden = True
            
            logger.info(f"已隱藏第7行到第539行（共{539-7+1}行）")
            
        except Exception as e:
            logger.error(f"隱藏行失敗: {e}")
            # 不拋出異常，讓程序繼續執行

def main():
    """主函數"""
    parser = argparse.ArgumentParser(description='將分類結果寫入Excel文件')
    parser.add_argument('results_file', help='分類結果JSON文件路徑')
    parser.add_argument('-o', '--output', help='輸出Excel文件路徑（可選）')
    parser.add_argument('-c', '--config', default='config.ini', help='配置文件路徑')
    
    args = parser.parse_args()
    
    print("分類結果寫入Excel工具")
    print("=" * 40)
    
    try:
        writer = ResultsToExcel(args.config)
        output_file = writer.process_results(args.results_file, args.output)
        
        print(f"\n✅ 處理完成！")
        print(f"📁 輸出文件: {output_file}")
        
    except Exception as e:
        logger.error(f"程序執行失敗: {e}")
        print(f"❌ 程序執行失敗: {e}")

if __name__ == "__main__":
    main()
